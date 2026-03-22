import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def read_excel(file) -> pd.DataFrame:
    """
    ExcelファイルをDataFrameに読み込む
    - header=0（1行目がヘッダー）
    - A列〜P列（インデックス0〜15）を対象
    - 列名はそのまま使用
    """
    filename = getattr(file, "name", "")
    if filename.endswith(".xls"):
        df = pd.read_excel(file, header=0, engine="xlrd")
    else:
        df = pd.read_excel(file, header=0, engine="openpyxl")

    # A〜P列（最大16列）のみ使用
    cols = df.columns.tolist()
    if len(cols) > 16:
        df = df.iloc[:, :16]

    return df


def write_excel(df: pd.DataFrame, original_filename: str) -> bytes:
    """
    チェック結果付きDataFrameをExcelバイト列として返す
    """
    output = io.BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # スタイル定義
    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(name="Arial", color="FFFFFF", bold=True)
    pass_fill = PatternFill(
        start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
    )
    fail_fill = PatternFill(
        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
    )
    normal_font = Font(name="Arial")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    nowrap_alignment = Alignment(vertical="top")

    # 列幅設定（A〜R）
    col_widths = {
        1: 15,   # A: LNO
        2: 15,   # B: TITLE
        3: 10,   # C: P
        4: 40,   # D: 日本語
        5: 40,   # E: AI訳文
        6: 30,   # F: 翻訳手順
        7: 12,   # G: X1
        8: 12,   # H: X2
        9: 12,   # I: X3
        10: 12,  # J: X4
        11: 12,  # K: MetricX
        12: 10,  # L: 評価①
        13: 10,  # M: 評価②
        14: 50,  # N: 評価③
        15: 50,  # O: 評価④
        16: 50,  # P: 評価⑤
        17: 15,  # Q: 品質チェック結果
        18: 50,  # R: 推奨英訳
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 折り返し対象列（D=4, E=5, N=14, O=15, R=18）
    wrap_cols = {4, 5, 14, 15, 18}

    total_cols = ws.max_column

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        # Q列（17列目）の値を取得してPASS/FAIL判定
        q_value = ""
        if total_cols >= 17:
            q_cell = ws.cell(row=row_idx, column=17)
            q_value = str(q_cell.value or "").strip().upper()

        for cell in row:
            col_num = cell.column

            if row_idx == 1:
                # ヘッダー行
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=False, vertical="center")
            else:
                # データ行
                cell.font = normal_font
                if col_num in wrap_cols:
                    cell.alignment = wrap_alignment
                else:
                    cell.alignment = nowrap_alignment

                # PASS/FAIL 行の背景色
                if q_value == "PASS":
                    cell.fill = pass_fill
                elif q_value == "FAIL":
                    cell.fill = fail_fill

    result_bytes = io.BytesIO()
    wb.save(result_bytes)
    result_bytes.seek(0)
    return result_bytes.getvalue()
